"""Allows the user to combine Excel workbooks.
Note that the modules used to do allow this by default,
as such, the results are not always perfect.
Simpler workbooks should pose no issue though."""

from tkinter import Tk
from tkinter.messagebox import showerror
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfilename
from copy import copy
from os import remove, system
from os.path import isfile
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image


def create_program_mainloop(transparent_icon_location = 'icon.ico') -> Tk:
    """Creates an instance of tk.Tk and replaces its icon with a transparent one."""

    if not isfile(transparent_icon_location):
        transparent_icon = Image.new('RGBA', (16, 16), (0, 0, 0, 0))
        transparent_icon.save(transparent_icon_location, 'ICO')

    root = Tk()
    root.title('')
    # hides its window
    root.withdraw()
    root.iconbitmap(True, transparent_icon_location)

    return root


def copy_sheet(source_sheet: Worksheet, target_sheet: Worksheet, simple_copy: bool):
    """Copies a sheet from one workbook to another, cell by cell."""

    copy_cells(source_sheet, target_sheet, simple_copy)
    copy_sheet_attributes(source_sheet, target_sheet)
    copy_row_dimensions(source_sheet, target_sheet)
    copy_column_dimensions(source_sheet, target_sheet)


def copy_cells(source_sheet: Worksheet, target_sheet: Worksheet, simple_copy: bool):
    """Copies cell contents as well as its properties from source to target sheet."""

    source_cell: Cell
    for (row, col), source_cell in source_sheet._cells.items():  #pylint: disable=protected-access
        target_cell: Cell = target_sheet.cell(column=col, row=row)

        # copies cell content
        target_cell._value = source_cell._value  #pylint: disable=protected-access
        target_cell.data_type = source_cell.data_type
        target_cell.number_format = copy(source_cell.number_format)

        if not simple_copy:
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)  #pylint: disable=protected-access

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


def copy_sheet_attributes(source_sheet: Worksheet, target_sheet: Worksheet):
    """Copies sheet attributes from source to target sheet."""

    target_sheet.auto_filter = copy(source_sheet.auto_filter)
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.print_options = copy(source_sheet.print_options)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)


def copy_row_dimensions(source_sheet: Worksheet, target_sheet: Worksheet):
    """Copies row hights from source to target sheet.
    Note: this is not always perfect as auto-fit has not been added as of yet
    Since openpyxl does not seem to include it."""

    for row in range(source_sheet.max_row + 1):
        target_sheet.row_dimensions[row].height = (
            source_sheet.row_dimensions[row].height)

def copy_column_dimensions(source_sheet: Worksheet, target_sheet: Worksheet):
    """Copies column widths from source to target sheet."""
    for col_index in range(source_sheet.max_column + 1)[1:]:
        target_col = target_sheet.column_dimensions[get_column_letter(col_index)]

        if col_index == 1:
            source_col = source_sheet.column_dimensions['A']
            target_col = target_sheet.column_dimensions['A']
            target_col.width = source_col.width

        else:
            index = col_index

            while index > 1:
                letter = get_column_letter(index)
                source_col = source_sheet.column_dimensions[letter]

                # it has been observed that columns that are next to each other
                # and of the same width were reverted to the default width, this fixes that issue
                if source_col.width == 13 and source_col.auto_size is False:
                    index -= 1
                else:
                    letter = get_column_letter(index)
                    source_col = source_sheet.column_dimensions[letter]
                    target_col.width = source_col.width
                    break
            else:
                target_col.width = source_sheet.column_dimensions['A'].width


def browse_excel_file(title="Choose a file") -> str:
    """Prompts user to select .xls or preferably .xlsx file."""

    file = askopenfilename(title=title,
                       filetypes=[("Excel (.xlsx) file", "*.xlsx"),
                                  ("Excel (.xls) file", "*.xls")])

    # program shuts down if no file has been selected
    if not file:
        sys.exit()

    return file


def browse_excel_files(title="Choose files") -> tuple[str, ...]:
    """Prompts user to select files for checking.
    Returns the list of files and their directory."""

    files = askopenfilenames(title=title,
                             filetypes=[("Excel (.xlsx) file", "*.xlsx"),
                                        ("Excel (.xls) file", "*.xls")])

    # program shuts down if no files have been selected
    if not files:
        sys.exit()

    return files


def save_excel(file: Workbook) -> str:
    """Saves the combined workbook."""
    file_saved = False
    while not file_saved:
        try:
            file_name = asksaveasfilename(defaultextension=".xlsx",
                                          filetypes=[("Excel file (.xlsx)",
                                                      "*.xlsx")])
            file_saved = True
        except PermissionError:
            showerror(
                title="Error occurred!",
                message=
                ("File could not be saved because it is already opened by another process "
                 "(likely Excel.exe). Please close it before continuing."))

    if not file_name:
        sys.exit()

    file.save(file_name)

    return file_name


def combine_workbooks(simple_copy: bool = False):
    """Combines two workbooks by copying sheets.
    This is done CELL by CELL and can take a little bit for larger workbooks.\n
    IMPORTANT: for large yet simple files such as CSVs and Datasets,
    please use the "simply_copy" mode.
    Performance increase is around 80-90%.\n
    Note, however, that this does not copy Fonts, Borders, Cell Fills, Cell alignment, etc.,
    so for the files that require this, please do not use this mode."""

    create_program_mainloop()

    source_files = browse_excel_files(title="Please select the source workbook")

    target_file = browse_excel_file(title="Please select the target workbook")


    target_workbook: Workbook = load_workbook(target_file)

    for source_file in source_files:
        source_workbook: Workbook = load_workbook(source_file)

        for sheet_name in source_workbook.sheetnames:
            source_sheet = source_workbook[sheet_name]
            target_sheet = target_workbook.create_sheet(sheet_name)
            copy_sheet(source_sheet, target_sheet, simple_copy)

    temp_file_name = 'temp_file.xlsx'
    target_workbook.save(temp_file_name)
    file = load_workbook(temp_file_name, read_only=False)
    remove(temp_file_name)

    file_saved = False
    while not file_saved:
        try:
            file_name = asksaveasfilename(defaultextension=".xlsx",
                                          filetypes=[("Excel file (.xlsx)",
                                                      "*.xlsx")])
            file_saved = True
        except PermissionError:
            showerror(
                title="Error occurred!",
                message=
                ("File could not be saved because it is already opened by another process "
                 "(likely Excel.exe). Please close it before continuing."))

    if not file_name:
        sys.exit()

    file.save(file_name)

    # opens the saved workbook
    system(f'"{file_name}"')


if __name__ == '__main__':
    combine_workbooks(simple_copy = False)
